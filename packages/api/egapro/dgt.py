"""DGT specific utils"""

import re
from datetime import date

import arrow
from naf import DB as NAF
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from progressist import ProgressBar

from egapro import config, constants, db, models
from egapro.schema import SCHEMA
from egapro.utils import flatten, remove_one_year


AGES = {
    ":29": "30",
    "30:39": "30-39",
    "40:49": "40-49",
    "50:": "50",
}
EFFECTIF = {"50:250": "50 à 250", "251:999": "251 à 999", "1000:": "1000 et plus"}


def truthy(val):
    return False if val is False else True


def isodatetime(val):
    if not val:
        return None
    # Excel doesn't know nothing about timezone, so let's transpose.
    when = arrow.get(val).to("Europe/Paris").naive
    return when


def isodate(val):
    if not val:
        return None
    return date.fromisoformat(val)


def code_naf(code):
    if not code:
        return None
    return f"{code} - {NAF[code]}"


def value_or_nc(value):
    return value if value is not None else "nc"


async def get_headers_columns():
    """Return a tuple of lists of (header_names, column_names) that we want in the export."""
    try:
        num_coefficient = await db.declaration.fetchval(
            "SELECT "
            "jsonb_array_length(data->'indicateurs'->'rémunérations'->'catégories') AS length "
            "FROM declaration WHERE data->'indicateurs'->'rémunérations' ? 'catégories' "
            "ORDER BY length DESC LIMIT 1;"
        )
    except db.NoData:
        num_coefficient = 0
    interesting_cols = (
        [
            ("source", "source"),
            ("URL_declaration", "URL_declaration"),  # Built from /data/id, see below
            ("Date_declaration", "déclaration.date", isodatetime),
            ("Date_modification", "modified_at", isodatetime),
            ("Email_declarant", "déclarant.email"),
            ("Nom", "déclarant.nom"),
            ("Prenom", "déclarant.prénom"),
            ("Telephone", "déclarant.téléphone"),
            ("Region", "entreprise.région", constants.REGIONS.get),
            ("Departement", "entreprise.département", constants.DEPARTEMENTS.get),
            ("Adresse", "entreprise.adresse"),
            ("CP", "entreprise.code_postal"),
            ("Commune", "entreprise.commune"),
            ("Pays", "entreprise.code_pays", constants.PAYS_ISO_TO_LIB.get),
            (
                "Annee_indicateurs",
                "déclaration.année_indicateurs",
            ),
            ("Periode_12mois", "déclaration.période_suffisante", truthy),
            ("Date_debut_periode", "déclaration.début_période_référence"),
            ("Date_fin_periode", "déclaration.fin_période_référence"),
            ("Structure", "entreprise.structure"),
            ("Tranche_effectif", "entreprise.effectif.tranche", EFFECTIF.get),
            ("Nb_salaries", "entreprise.effectif.total"),
            ("Nom_Entreprise", "entreprise.raison_sociale"),
            ("SIREN", "entreprise.siren"),
            ("Code_NAF", "entreprise.code_naf", code_naf),
            ("Nom_UES", "entreprise.ues.nom"),
            # Inclure entreprise déclarante
            ("Nb_ets_UES", "entreprise.nombre_ues"),
            (
                "Date_publication",
                "déclaration.publication.date",
                isodate,
            ),
            ("Site_internet_publication", "déclaration.publication.url"),
            ("Modalites_publication", "déclaration.publication.modalités"),
            ("Indic1_calculable", "indicateurs.rémunérations.non_calculable_bool"),
            (
                "Indic1_motif_non_calculable",
                "indicateurs.rémunérations.non_calculable",
            ),
            ("Indic1_modalite_calcul", "indicateurs.rémunérations.mode"),
            (
                "Indic1_date_consult_CSE",
                "indicateurs.rémunérations.date_consultation_cse",
                isodate,
            ),
            ("Indic1_nb_coef_niv", "indicateurs.rémunérations.len_categories"),
            ("Indic1_Ouv", "indicateurs.rémunérations.Indic1_Ouv"),
            ("Indic1_Emp", "indicateurs.rémunérations.Indic1_Emp"),
            ("Indic1_TAM", "indicateurs.rémunérations.Indic1_TAM"),
            ("Indic1_IC", "indicateurs.rémunérations.Indic1_IC"),
        ]
        + [
            (
                f"Indic1_Niv{index_coef}",
                f"indicateurs.rémunérations.catégories.{index_coef}",
            )
            for index_coef in range(num_coefficient)
        ]
        + [
            ("Indic1_resultat", "indicateurs.rémunérations.résultat"),
            (
                "Indic1_population_favorable",
                "indicateurs.rémunérations.population_favorable",
            ),
            ("Indic2_calculable", "indicateurs.augmentations.non_calculable_bool"),
            (
                "Indic2_motif_non_calculable",
                "indicateurs.augmentations.non_calculable",
            ),
        ]
        + [
            (
                f"Indic2_{CSP}",
                f"indicateurs.augmentations.catégories.{index_csp}",
            )
            for (index_csp, CSP) in enumerate(["Ouv", "Emp", "TAM", "IC"])
        ]
        + [
            ("Indic2_resultat", "indicateurs.augmentations.résultat"),
            (
                "Indic2_population_favorable",
                "indicateurs.augmentations.population_favorable",
            ),
            ("Indic3_calculable", "indicateurs.promotions.non_calculable_bool"),
            ("Indic3_motif_non_calculable", "indicateurs.promotions.non_calculable"),
        ]
        + [
            (
                f"Indic3_{CSP}",
                f"indicateurs.promotions.catégories.{index_csp}",
            )
            for (index_csp, CSP) in enumerate(["Ouv", "Emp", "TAM", "IC"])
        ]
        + [
            ("Indic3_resultat", "indicateurs.promotions.résultat"),
            (
                "Indic3_population_favorable",
                "indicateurs.promotions.population_favorable",
            ),
            (
                "Indic2et3_calculable",
                "indicateurs.augmentations_et_promotions.non_calculable_bool",
            ),
            (
                "Indic2et3_motif_non_calculable",
                "indicateurs.augmentations_et_promotions.non_calculable",
            ),
            (
                "Indic2et3_resultat_pourcent",
                "indicateurs.augmentations_et_promotions.résultat",
            ),
            (
                "Indic2et3_resultat_nb_sal",
                "indicateurs.augmentations_et_promotions.résultat_nombre_salariés",
            ),
            (
                "Indic2et3_population_favorable",
                "indicateurs.augmentations_et_promotions.population_favorable",
            ),
            ("Indic4_calculable", "indicateurs.congés_maternité.non_calculable_bool"),
            (
                "Indic4_motif_non_calculable",
                "indicateurs.congés_maternité.non_calculable",
            ),
            ("Indic4_resultat", "indicateurs.congés_maternité.résultat"),
            ("Indic5_resultat", "indicateurs.hautes_rémunérations.résultat"),
            (
                "Indic5_sexe_sur_represente",
                "indicateurs.hautes_rémunérations.population_favorable",
            ),
            ("Indicateur_1", "indicateurs.rémunérations.note"),
            (
                "Indicateur_1_objectif",
                "indicateurs.rémunérations.objectif_de_progression",
            ),
            ("Indicateur_2", "indicateurs.augmentations.note"),
            (
                "Indicateur_2_objectif",
                "indicateurs.augmentations.objectif_de_progression",
            ),
            ("Indicateur_3", "indicateurs.promotions.note"),
            ("Indicateur_3_objectif", "indicateurs.promotions.objectif_de_progression"),
            ("Indicateur_2et3", "indicateurs.augmentations_et_promotions.note"),
            (
                "Indicateur_2et3_objectif",
                "indicateurs.augmentations_et_promotions.objectif_de_progression",
            ),
            (
                "Indicateur_2et3_PourCent",
                "indicateurs.augmentations_et_promotions.note_en_pourcentage",
            ),
            (
                "Indicateur_2et3_ParSal",
                "indicateurs.augmentations_et_promotions.note_nombre_salariés",
            ),
            ("Indicateur_4", "indicateurs.congés_maternité.note"),
            (
                "Indicateur_4_objectif",
                "indicateurs.congés_maternité.objectif_de_progression",
            ),
            ("Indicateur_5", "indicateurs.hautes_rémunérations.note"),
            (
                "Indicateur_5_objectif",
                "indicateurs.hautes_rémunérations.objectif_de_progression",
            ),
            ("Nombre_total_points obtenus", "déclaration.points"),
            (
                "Nombre_total_points_pouvant_etre_obtenus",
                "déclaration.points_calculables",
            ),
            ("Resultat_final_sur_100_points", "déclaration.index", value_or_nc),
            ("Mesures_correction", "déclaration.mesures_correctives"),
            (
                "Modalites_objectifs_mesure",
                "déclaration.publication.modalités_objectifs_mesures",
            ),
            (
                "Date_publication_mesures",
                "déclaration.publication.date_publication_mesures",
            ),
            (
                "Date_publication_objectifs",
                "déclaration.publication.date_publication_objectifs",
            ),
            ("Plan_relance", "entreprise.plan_relance"),
        ]
    )
    headers = []
    columns = []
    for header, column, *fmt in interesting_cols:
        headers.append(header)
        columns.append((column, fmt[0] if fmt else lambda x: x))
    return (headers, columns)


WHITE_SPACES = re.compile(r"\s+")


def clean_cell(value):
    if isinstance(value, str):
        value = WHITE_SPACES.sub(" ", ILLEGAL_CHARACTERS_RE.sub("", value).strip())
    return value


async def as_xlsx(max_rows=None, debug=False):
    """Export des données au format souhaité par la DGT.

    :max_rows:          Max number of rows to process.
    :debug:             Turn on debug to be able to read the generated Workbook
    """
    print("Reading from DB")
    records = await db.declaration.completed()
    print("Flattening JSON")
    if max_rows:
        records = records[:max_rows]
    wb = Workbook(write_only=not debug)
    ws = wb.create_sheet()
    ws.title = "BDD REPONDANTS"
    wb.active = ws
    ws_ues = wb.create_sheet()
    ws_ues.title = "BDD UES détail entreprises"
    ws_ues.append(
        [
            "Annee_indicateurs",
            "Region",
            "Departement",
            "Adresse",
            "CP",
            "Commune",
            "Tranche_effectif",
            "Nom_UES",
            "Siren_entreprise_declarante",
            "Nom_entreprise_declarante",
            "Nom_entreprise",
            "Siren",
        ]
    )
    headers, columns = await get_headers_columns()
    ws.append(headers)
    bar = ProgressBar(prefix="Computing", total=len(records))
    for record in bar.iter(records):
        data = record.data
        if not data:
            continue
        ues_data(ws_ues, data)
        data = prepare_record(data)
        data["modified_at"] = record["modified_at"]
        ws.append([clean_cell(fmt(data.get(c))) for c, fmt in columns])
    return wb


def ues_data(sheet, data):
    data = models.Data(data)
    entreprises = data.path("entreprise.ues.entreprises")
    if not entreprises:
        return
    region = constants.REGIONS.get(data.path("entreprise.région"))
    departement = constants.DEPARTEMENTS.get(data.path("entreprise.département"))
    adresse = data.path("entreprise.adresse")
    cp = data.path("entreprise.code_postal")
    commune = data.path("entreprise.commune")
    tranche = EFFECTIF.get(data.path("entreprise.effectif.tranche"))
    nom = data.path("entreprise.ues.nom")
    rows = [
        [
            data.year,
            region,
            departement,
            adresse,
            cp,
            commune,
            tranche,
            nom,
            data.siren,
            data.company,
            data.path("entreprise.raison_sociale"),
            data.siren,
        ]
    ]
    for ues in entreprises or []:
        rows.append(
            [
                data.year,
                region,
                departement,
                adresse,
                cp,
                commune,
                tranche,
                nom,
                data.siren,
                data.company,
                ues["raison_sociale"],
                ues["siren"],
            ]
        )
    for row in rows:
        sheet.append(clean_cell(cell) for cell in row)

# => "indicateur": (<250, 250+)
indicateur_nc_or_not: dict[str, tuple] = {
    "rémunérations": ("nc", "nc"),
    "augmentations": ("", "nc"),
    "promotions": ("", "nc"),
    "augmentations_et_promotions": ("nc", ""),
    "congés_maternité": ("nc", "nc"),
    "hautes_rémunérations": ("nc", "nc"),
}

def prepare_record(data: models.Data):

    # Before flattening.
    data["URL_declaration"] = f"'{config.DOMAIN}{data.uri}"
    effectif = data["entreprise"]["effectif"]["tranche"]
    prepare_entreprise(data["entreprise"])
    prepare_declaration(data["déclaration"])
    periode_suffisante = data["déclaration"].get("période_suffisante") is not False
    if not periode_suffisante:
        data.setdefault("indicateurs", {})
        for key, (value_eff_lt_250, value_eff_250_999) in indicateur_nc_or_not.items():
            if key not in data["indicateurs"]:
                data["indicateurs"][key] = {}
            data["indicateurs"][key]["note"] = value_eff_lt_250 if effectif == "50:250" else value_eff_250_999
    elif "indicateurs" in data:
        prepare_remunerations(data["indicateurs"]["rémunérations"])
        prepare_conges_maternite(data["indicateurs"]["congés_maternité"])
        if effectif == "50:250":
            prepare_augmentations_et_promotions(
                data["indicateurs"]["augmentations_et_promotions"]
            )
        else:
            prepare_augmentations(data["indicateurs"]["augmentations"])
            prepare_promotions(data["indicateurs"]["promotions"])

    return flatten(data, flatten_lists=True)


def prepare_entreprise(data):
    nombre_ues = len(data.get("ues", {}).get("entreprises", []))
    data["structure"] = (
        "Unité Economique et Sociale (UES)" if nombre_ues else "Entreprise"
    )
    data["nombre_ues"] = nombre_ues + 1 if nombre_ues else None


def prepare_declaration(data):
    fin_periode_reference = data.get("fin_période_référence")
    if fin_periode_reference:
        data["début_période_référence"] = remove_one_year(
            date.fromisoformat(fin_periode_reference)
        )
        data["fin_période_référence"] = date.fromisoformat(fin_periode_reference)


def prepare_remunerations(data):
    try:
        indic1_categories = data["catégories"]
    except KeyError:
        indic1_categories = []
    indic1_nv_niveaux = len(indic1_categories) or None
    indic1_mode = data.get("mode")
    data["len_categories"] = indic1_nv_niveaux if indic1_mode != "csp" else None
    calculable = not data.get("non_calculable")
    if calculable:
        # DGT want data to be in different columns whether its csp or any coef.
        csp_names = ["Ouv", "Emp", "TAM", "IC"]
        for idx, category in enumerate(indic1_categories):
            tranches = category.get("tranches", {})
            key = f"catégories.{idx}"
            if indic1_mode == "csp":
                key = f"Indic1_{csp_names[idx]}"
            values = [
                tranches.get(":29"),
                tranches.get("30:39"),
                tranches.get("40:49"),
                tranches.get("50:"),
            ]
            # Prevent "-0.0" or "0.0" or "12.0" as str representation
            values = [int(v) if v is not None and v % 1 == 0 else v for v in values]
            values = [str(round(v, 2) + 0) if v is not None else "nc" for v in values]
            data[key] = ";".join(values)
    else:
        data["note"] = "nc"
    data["non_calculable_bool"] = calculable


def prepare_promotions(data):
    calculable = not data.get("non_calculable")
    data["non_calculable_bool"] = calculable
    if not calculable:
        data["note"] = "nc"


def prepare_augmentations(data):
    calculable = not data.get("non_calculable")
    data["non_calculable_bool"] = calculable
    if not calculable:
        data["note"] = "nc"


def prepare_augmentations_et_promotions(data):
    calculable = not data.get("non_calculable")
    data["non_calculable_bool"] = calculable
    if not calculable:
        data["note"] = "nc"
        data["note_en_pourcentage"] = "nc"
        data["note_nombre_salariés"] = "nc"


def prepare_conges_maternite(data):
    calculable = not data.get("non_calculable")
    data["non_calculable_bool"] = calculable
    if not calculable:
        data["note"] = "nc"
